import openpyxl

def findMatch(state,district,county):
    state = state.lower()
    district = district.lower()
    county = county.lower()
    print(state, district, county)
    path = "book1.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row):
        reqLoc=[]
        s = sheet.cell(row=i, column=2).value.strip().lower()
        d = sheet.cell(row=i, column=3).value.strip().lower()
        c = sheet.cell(row=i, column=4).value.strip().lower()
        print(c, 'and', county)
        # if state==s and district==d and county==c:
        #     print("found state and district.")
        #     break
        if state==s and d==district and county==c:
            print("found")
            for j in range(1, 24):
                cell = sheet.cell(row=i,column=j)
                reqLoc.append(cell.value)
                print(cell.value)
            break
    print(reqLoc)
    return reqLoc


# st = 'Assam'
# di = 'Baksa'
# co = 'Simla'
# print(findMatch(st,di,co))
